#!/usr/bin/env python3
"""First-pass annual costings of four UK trade-adjustment-assistance (TAA)
worker-side policies from public data.

One rerunnable script, mirroring the conventions of
https://github.com/PolicyEngine/uk-trade-shock-study : pinned raw inputs
(verified by SHA256), all reported numbers generated (never typed) into
LaTeX macros, and every assumption labelled in NOTES.md.

Inputs (data/raw/, see NOTES.md for URLs and provenance):
  - red02aug2026.xlsx                  ONS LFS RED02, redundancies by industry
  - red01nsaaug2026.xls                ONS LFS RED01 (NSA), redundancy levels
                                       people/men/women back to 1995
  - ashetable42025provisional.zip      ASHE 2025 provisional, Table 4 (unzipped
                                       to data/raw/ashe_table4/)
  - benefit_pension_rates_2026_27.pdf  DWP proposed benefit rates 2026-27
  - pathways_to_work_green_paper.html  DWP Green Paper (UI proposal)
  - port_talbot_22m_press_release.html gov.uk (Port Talbot £122m)
  - nao_british_steel_press_release.html NAO (Scunthorpe £377m)

Outputs: out/results.json, out/generated_taa.tex

Usage:  .venv/bin/python costing.py     (deps: openpyxl + xlrd)
"""

import hashlib
import json
import re
import statistics
from pathlib import Path

import openpyxl
import xlrd

ROOT = Path(__file__).resolve().parent
RAW = ROOT / "data" / "raw"
OUT = ROOT / "out"

# --------------------------------------------------------------------------
# 0. Pinned raw inputs (SHA256 recorded at download time, 18 Aug 2026)
# --------------------------------------------------------------------------
PINNED = {
    "red02aug2026.xlsx":
        "f653c80c7f48d9f9becffd0002d616ba4b524711ed2f390186b0c242d193feb3",
    "ashetable42025provisional.zip":
        "581d23bb7e1746032246b720efe82a34dfb9708f0407b4d752e654d8cecb1284",
    "benefit_pension_rates_2026_27.pdf":
        "c13111a6a4a939778bce54b248ce525470b98b730276715632fd5d1eaec3c019",
    "pathways_to_work_green_paper.html":
        "0af072c178ca3f601f198e5677708f7d5ff7861a3f64b51df5ad0127b7f5a073",
    "port_talbot_22m_press_release.html":
        "cb845995ce8d93eecd8d9cb15357d7a74c4d7b49c28166e382cd3216bb581fd9",
    "nao_british_steel_press_release.html":
        "2f0f02494e55f694eefea6b9b66542209fd98ec7f0c075eede32d730668345e8",
    # Added 18 Aug 2026 for the cyclical-stress and demographic modules:
    "red01nsaaug2026.xls":
        "79a52829a002dc0942e2249839056c5514ceb4998ff0f6f7dd9591ad22e75c2d",
}


def verify_inputs():
    for name, expected in PINNED.items():
        p = RAW / name
        if not p.exists():
            raise FileNotFoundError(f"Missing pinned input {p}; see NOTES.md")
        got = hashlib.sha256(p.read_bytes()).hexdigest()
        if got != expected:
            raise ValueError(f"Hash mismatch for {name}: {got} != {expected}")
    print(f"Verified {len(PINNED)} pinned raw inputs.")


# --------------------------------------------------------------------------
# 1. Redundancy flows — ONS RED02 (LFS), 'Industry - levels' sheet
#    Annual flow = sum of the latest four NON-overlapping rolling-quarter
#    periods (the sheet publishes overlapping 3-month rolling windows).
# --------------------------------------------------------------------------
RED02_QUARTERS = ["Jul-Sep 2025", "Oct-Dec 2025", "Jan-Mar 2026", "Apr-Jun 2026"]
RED02_YEAR_LABEL = "Jul 2025 - Jun 2026"


def read_red02():
    wb = openpyxl.load_workbook(RAW / "red02aug2026.xlsx", read_only=True)
    ws = wb["Industry - levels"]
    rows = list(ws.iter_rows(values_only=True))
    codes = rows[6]  # header row of SIC section codes: A-U, ..., C, ...
    col_all = codes.index("A-U")
    col_manuf = codes.index("C")
    data = {}
    for r in rows:
        if isinstance(r[0], str) and r[0].strip() in RED02_QUARTERS:
            data[r[0].strip()] = (float(r[col_all]), float(r[col_manuf]))
    missing = [q for q in RED02_QUARTERS if q not in data]
    if missing:
        raise ValueError(f"RED02 quarters not found: {missing}")
    flow_all = sum(data[q][0] for q in RED02_QUARTERS)
    flow_manuf = sum(data[q][1] for q in RED02_QUARTERS)
    return flow_all, flow_manuf, {q: data[q] for q in RED02_QUARTERS}


# --------------------------------------------------------------------------
# 2. Earnings — ASHE 2025 provisional Table 4.1a (gross weekly pay),
#    'Full-Time' sheet.  Columns: Description, Code, jobs(000s), Median,
#    %chg, Mean, %chg, p10, p20, p25, ...
# --------------------------------------------------------------------------
ASHE_FILE = ("ashe_table4/PROV - SIC07 Industry (2) SIC2007 Table 4.1a   "
             "Weekly pay - Gross 2025.xlsx")


def read_ashe():
    wb = openpyxl.load_workbook(RAW / ASHE_FILE, read_only=True)
    ws = wb["Full-Time"]
    out = {}
    for r in ws.iter_rows(values_only=True, max_col=10):
        desc = str(r[0]).strip() if r[0] else ""
        if desc == "ALL EMPLOYEES":
            out["all"] = dict(median=float(r[3]), mean=float(r[5]),
                              p25=float(r[9]))
        elif desc == "MANUFACTURING":
            out["manuf"] = dict(median=float(r[3]), mean=float(r[5]),
                                p25=float(r[9]))
    if set(out) != {"all", "manuf"}:
        raise ValueError(f"ASHE rows not found; got {set(out)}")
    return out


# --------------------------------------------------------------------------
# 2b. Historical redundancy series and demographic composition
#     - RED01 (NSA): whole-economy levels, people/men/women, monthly-rolling
#       windows Mar-May 1995 onwards ('..' before then).  Annual flow for a
#       calendar year = sum of the four non-overlapping quarters
#       Jan-Mar + Apr-Jun + Jul-Sep + Oct-Dec (same convention as section 1).
#       Some period labels carry a trailing footnote digit (e.g.
#       'Jan-Mar 20193' = Jan-Mar 2019, footnote 3) -- stripped by regex.
#     - RED02 'Industry - levels': same annual sums for all-industry and
#       manufacturing, available from calendar 2009.
#     - RED02 'Age - levels': redundancies by age band (16-24/25-34/35-49/
#       50+), from Jan-Mar 2009.
# --------------------------------------------------------------------------
CAL_QUARTERS = ["Jan-Mar", "Apr-Jun", "Jul-Sep", "Oct-Dec"]
RED01_PERIOD_RE = re.compile(
    r"^(Jan-Mar|Apr-Jun|Jul-Sep|Oct-Dec) (\d{4})\d?$")  # optional footnote


def read_red01_annual():
    """Calendar-year annual redundancy flows (people, men, women) from RED01
    NSA, plus the latest-year (Jul 2025-Jun 2026) people/men totals for the
    sex composition.  Returns (annual, latest_year_sex)."""
    wb = xlrd.open_workbook(RAW / "red01nsaaug2026.xls")
    ws = wb.sheet_by_name("All")
    by_year = {}
    latest = {"people": 0.0, "men": 0.0}
    for i in range(ws.nrows):
        lab = str(ws.cell_value(i, 0)).strip()
        people, men, women = (ws.cell_value(i, 1), ws.cell_value(i, 3),
                              ws.cell_value(i, 5))
        if not isinstance(people, float):
            continue
        m = RED01_PERIOD_RE.match(lab)
        if m and 1995 <= int(m.group(2)) <= 2026:
            y = int(m.group(2))
            by_year.setdefault(y, []).append((people, men, women))
        # latest-year sex composition uses the same four non-overlapping
        # quarters as the section-1 eligible flow (Jul 2025 - Jun 2026):
        if lab in RED02_QUARTERS:
            latest["people"] += people
            latest["men"] += men
    annual = {y: {"people": sum(v[0] for v in q),
                  "men": sum(v[1] for v in q),
                  "women": sum(v[2] for v in q)}
              for y, q in by_year.items() if len(q) == 4}  # complete years
    if not annual or latest["people"] == 0:
        raise ValueError("RED01 parse failed")
    return annual, latest


def read_red02_history_and_age():
    """From the pinned RED02: (a) calendar-year annual all-industry and
    manufacturing flows (available 2009 onwards); (b) latest-year
    (Jul 2025 - Jun 2026) age composition of all-industry redundancies."""
    wb = openpyxl.load_workbook(RAW / "red02aug2026.xlsx", read_only=True)
    ws = wb["Industry - levels"]
    rows = list(ws.iter_rows(values_only=True))
    codes = rows[6]
    col_all, col_manuf = codes.index("A-U"), codes.index("C")
    period = {}
    for r in rows:
        if (isinstance(r[0], str)
                and isinstance(r[col_all], (int, float))
                and isinstance(r[col_manuf], (int, float))):
            period[r[0].strip()] = (float(r[col_all]), float(r[col_manuf]))
    annual = {}
    for y in range(2009, 2027):
        labs = [f"{q} {y}" for q in CAL_QUARTERS]
        if all(lab in period for lab in labs):
            annual[y] = {"all": sum(period[lab][0] for lab in labs),
                         "manuf": sum(period[lab][1] for lab in labs)}
    ws_age = wb["Age - levels"]
    rows_age = list(ws_age.iter_rows(values_only=True))
    # columns: 0 period, 1 all 16+, 2 16-24, 3 25-34, 4 35-49, 5 50+
    age = {"all": 0.0, "fifty_plus": 0.0}
    for r in rows_age:
        if isinstance(r[0], str) and r[0].strip() in RED02_QUARTERS:
            age["all"] += float(r[1])
            age["fifty_plus"] += float(r[5])
    if not annual or age["all"] == 0:
        raise ValueError("RED02 history/age parse failed")
    return annual, age


# --------------------------------------------------------------------------
# 3. Statutory / policy parameters (sources: pinned PDF + Green Paper HTML;
#    every figure cross-checked against the raw file — see NOTES.md)
# --------------------------------------------------------------------------
PARAMS = {
    # DWP 'Benefit and pension rates 2026-2027' (pinned PDF):
    "jsa_weekly_2026_27": 95.55,        # NS/contribution-based JSA, 25+
    "jsa_duration_weeks": 26,           # statutory NS JSA maximum
    "uc_single_25plus_monthly_2026_27": 424.90,
    "esa_pa_25plus_2026_27": 95.55,     # NS ESA personal allowance 25+
    "esa_support_component_2026_27": 50.35,
    # Pathways to Work Green Paper (pinned HTML): UI paid at "the current
    # higher rate of NS ESA" (= £140.55/wk at 2025-26 rates); we cost at the
    # 2026-27 uprated equivalent PA + support component for consistency:
    "ui_weekly_2026_27": 95.55 + 50.35,  # = 145.90
    "ui_duration_months_options": [6, 9, 12],  # consulted range "e.g. 6-12"
    # Imported calibrations (companion study FRS calibration; NOT
    # re-estimated here): re-employment wage penalty 28.3% all-employee,
    # 14.0% hours-controlled -> low/high scenario values 15% / 28%,
    # central = midpoint (ASSUMPTION):
    "penalty": {"low": 0.15, "central": 0.215, "high": 0.28},
    # Re-employment within a year (ASSUMPTION bracketing UK displacement
    # literature; see NOTES.md):
    "reemployment": {"low": 0.50, "central": 0.70, "high": 0.85},
    # Wage-insurance design:
    "wi_replacement": 0.50, "wi_duration_years": 2.0,
    # UI arm: contribution-condition share (ASSUMPTION):
    "contrib_share": {"low": 0.80, "central": 0.90, "high": 1.00},
    "ui_duration_weeks": {"low": 26, "central": 39, "high": 52},  # 6/9/12 mo
    # Retraining stipend (ASSUMPTION, brackets US TAA training take-up):
    "train_stipend": 3000.0,
    "train_takeup": {"low": 0.20, "central": 0.40, "high": 0.60},
    # UC capital-disregard placeholder (SPECULATIVE):
    "uc_months": 6,
    "uc_capital_gated_share": {"low": 0.20, "central": 0.35, "high": 0.50},
    # Trade-exposed manufacturing memo share (ASSUMPTION):
    "exposed_share_of_manuf": 0.30,
    # Cross-check anchor (pinned press releases): ad-hoc steel support
    "steel_port_talbot_m": 102.0,    # HMG share only (Transition Board totals £122m incl. £20m Tata)
    "steel_scunthorpe_m": 377.0,     # NAO, Apr 2025 - Jan 2026
    # ---- Module 1: wage-insurance parameter grid (narrow, central take-up) --
    "wi_grid_replacement": [0.30, 0.50, 0.70],
    "wi_grid_duration_years": [1, 2, 3],
    # penalty axis reuses PARAMS["penalty"] {15%, 21.5%, 28%}
    # ISERBS-parameterised cell (1995 civil-service redundancy insurance
    # analogue): top-up to 90% of previous pay for 78 weeks (1.5 years):
    "iserbs_target_replacement": 0.90,
    "iserbs_duration_weeks": 78,
    # ---- Module 2: cyclical stress episodes (calendar years) ----------------
    "stress_gfc_years": [2008, 2009],    # peak = max annual flow within
    "stress_covid_years": [2020, 2021],
    # ---- Module 3: self-financing break-even, 2026-27 tax/NIC (gov.uk) ------
    "income_tax_personal_allowance": 12570.0,  # frozen since 2021-22
    "income_tax_basic_rate": 0.20,
    "nic_employee_main_rate": 0.08,      # Class 1 main rate since 6 Apr 2024
    "nic_primary_threshold_annual": 12570.0,   # aligned with PA
}

SCEN = ["low", "central", "high"]


# --------------------------------------------------------------------------
# 4. Policy-arm costings (annual steady state, £m/yr)
# --------------------------------------------------------------------------
def cost_arms(flows, ashe):
    p = PARAMS
    # eligibility -> (annual flow, pre-displacement mean weekly earnings)
    elig = {
        "narrow":  (flows["manuf"], ashe["manuf"]["mean"]),
        "broad":   (flows["all"], ashe["all"]["mean"]),
        "exposed": (flows["manuf"] * p["exposed_share_of_manuf"],
                    ashe["manuf"]["mean"]),
    }
    res = {"arms": {}, "per_beneficiary": {}, "eligible_flows":
           {k: v[0] for k, v in elig.items()}}

    # --- A. Wage insurance -------------------------------------------------
    # 50% of the wage gap (penalty x pre-displacement mean weekly earnings),
    # paid to re-employed displaced workers for up to 2 years.  Steady-state
    # annual cost = flow x reemp x 0.5 x penalty x mean x 52 x duration(2),
    # i.e. two cohorts in payment at once, no wage catch-up (ASSUMPTION).
    A = {}
    for e, (flow, mean_wk) in elig.items():
        A[e] = {}
        for s in SCEN:
            per_worker_yr = (p["wi_replacement"] * p["penalty"][s]
                             * mean_wk * 52)
            cost = (flow * p["reemployment"][s] * per_worker_yr
                    * p["wi_duration_years"]) / 1e6
            A[e][s] = cost
    res["arms"]["wage_insurance"] = A
    res["per_beneficiary"]["wage_insurance_annual_payment"] = {
        s: {e: p["wi_replacement"] * p["penalty"][s] * elig[e][1] * 52
            for e in elig} for s in SCEN}

    # --- B. Unemployment Insurance extension -------------------------------
    # Green Paper UI (NS ESA higher rate, 2026-27 uprated = £145.90/wk) for
    # 6/9/12 months vs NS JSA baseline (£95.55/wk x 26wk).  Incremental cost
    # per claimant x flow x contribution-condition share.
    B = {}
    jsa_base = p["jsa_weekly_2026_27"] * p["jsa_duration_weeks"]
    for e, (flow, _) in elig.items():
        B[e] = {}
        for s in SCEN:
            ui_total = p["ui_weekly_2026_27"] * p["ui_duration_weeks"][s]
            incr = ui_total - jsa_base
            B[e][s] = flow * p["contrib_share"][s] * incr / 1e6
    res["arms"]["ui_extension"] = B
    res["per_beneficiary"]["ui_extension_incremental"] = {
        s: p["ui_weekly_2026_27"] * p["ui_duration_weeks"][s] - jsa_base
        for s in SCEN}

    # --- C. Retraining stipend ---------------------------------------------
    C = {}
    for e, (flow, _) in elig.items():
        C[e] = {s: flow * p["train_takeup"][s] * p["train_stipend"] / 1e6
                for s in SCEN}
    res["arms"]["retraining"] = C
    res["per_beneficiary"]["retraining_stipend"] = p["train_stipend"]

    # --- D. UC redundancy-pay capital disregard (PLACEHOLDER) --------------
    # MAXIMUM exposure bound only: flow x 6 months of the UC single-25+
    # standard allowance x share plausibly capital-gated.  SPECULATIVE —
    # requires the FRS microsimulation for a credible estimate.
    D = {}
    uc_6m = p["uc_single_25plus_monthly_2026_27"] * p["uc_months"]
    for e, (flow, _) in elig.items():
        D[e] = {s: flow * p["uc_capital_gated_share"][s] * uc_6m / 1e6
                for s in SCEN}
    res["arms"]["uc_capital_disregard_bound"] = D
    res["per_beneficiary"]["uc_six_month_standard_allowance"] = uc_6m

    # --- Cross-checks -------------------------------------------------------
    steel = p["steel_port_talbot_m"] + p["steel_scunthorpe_m"]
    totals = {}
    for e in elig:
        totals[e] = {s: sum(res["arms"][a][e][s] for a in res["arms"])
                     for s in SCEN}
    res["totals"] = totals
    res["cross_checks"] = {
        "steel_support_m": steel,
        "total_central_pct_of_steel": {
            e: 100 * totals[e]["central"] / steel for e in elig},
    }
    return res


# --------------------------------------------------------------------------
# 4b. Added analysis modules (grid, ISERBS, cyclical stress, break-even,
#     demographics).  All narrow eligibility (manufacturing flow, ASHE
#     manufacturing mean weekly pay), central take-up/re-employment.
# --------------------------------------------------------------------------
def analysis_modules(res, flows, ashe, red01_annual, red01_latest_sex,
                     red02_annual, red02_age):
    p = PARAMS
    flow = flows["manuf"]                      # narrow eligible flow /yr
    mean_wk = ashe["manuf"]["mean"]            # £/wk pre-displacement
    reemp = p["reemployment"]["central"]       # 0.70
    mods = {}

    # --- Module 1: wage-insurance parameter grid ---------------------------
    # cost(R, D, pen) = flow x reemp x R x pen x mean_wk x 52 x D / 1e6
    # (same steady-state convention as Arm A: D annual cohorts in payment,
    # no catch-up/attrition).
    grid = {}
    for pen_name in ["low", "central", "high"]:
        pen = p["penalty"][pen_name]
        for R in p["wi_grid_replacement"]:
            for D in p["wi_grid_duration_years"]:
                grid[f"pen{pen_name}_r{int(R*100)}_d{D}"] = (
                    flow * reemp * R * pen * mean_wk * 52 * D / 1e6)
    mods["wi_grid"] = {
        "cells_m_per_yr": grid,
        "min_m_per_yr": min(grid.values()),
        "max_m_per_yr": max(grid.values()),
        "convention": "narrow eligibility, central re-employment 0.70; "
                      "steady state = D cohorts in payment",
    }

    # --- Module 1b: ISERBS-parameterised cell ------------------------------
    # Payment = max(0, 90% of previous earnings - new earnings).  At the
    # central 21.5% penalty, new pay = 78.5% of previous, so the top-up rate
    # is 90 - 78.5 = 11.5% of previous pay; 78 weeks (1.5 years).
    topup_rate = max(0.0, p["iserbs_target_replacement"]
                     - (1 - p["penalty"]["central"]))
    iserbs_per_worker_yr = topup_rate * mean_wk * 52
    iserbs_cost = (flow * reemp * topup_rate * mean_wk
                   * p["iserbs_duration_weeks"]) / 1e6
    mods["iserbs"] = {
        "topup_rate_of_prior_pay": topup_rate,
        "per_worker_per_yr": iserbs_per_worker_yr,
        "cost_m_per_yr": iserbs_cost,
        "duration_weeks": p["iserbs_duration_weeks"],
    }

    # --- Module 2: cyclical/historical stress ------------------------------
    # Manufacturing annual flows in stress episodes.  GFC and Covid peaks:
    # direct RED02 manufacturing calendar-year sums (RED02 industry detail
    # starts Jan-Mar 2009; RED01 whole-economy confirms 2009 > 2008 and
    # 2020 > 2021, so both peak years lie inside RED02 coverage).
    # Historical median: RED01 whole-economy median annual flow over the
    # complete calendar years, scaled to manufacturing by the CURRENT
    # manufacturing share of all-industry redundancies (ASSUMPTION).
    hist_years = sorted(red01_annual)
    people = {y: red01_annual[y]["people"] for y in hist_years}
    gfc_year = max(p["stress_gfc_years"], key=lambda y: people[y])
    covid_year = max(p["stress_covid_years"], key=lambda y: people[y])
    manuf_share_now = flows["manuf"] / flows["all"]
    med_all = statistics.median(people.values())
    stress_flows = {
        "gfc": red02_annual[gfc_year]["manuf"],
        "covid": red02_annual[covid_year]["manuf"],
        "hist_median": med_all * manuf_share_now,
    }
    total_central_narrow = res["totals"]["narrow"]["central"]
    # All four arms are linear in the eligible flow, so the stressed total
    # programme cost is the central narrow total scaled by the flow ratio.
    stress_costs = {k: total_central_narrow * v / flow
                    for k, v in stress_flows.items()}
    mods["stress"] = {
        "red01_annual_people": {str(y): people[y] for y in hist_years},
        "hist_median_all_industry": med_all,
        "hist_years_range": [hist_years[0], hist_years[-1]],
        "gfc_peak_year": gfc_year, "covid_peak_year": covid_year,
        "manuf_share_of_all_current": manuf_share_now,
        "stress_manuf_flows": stress_flows,
        "flow_ratio_to_current": {k: v / flow
                                  for k, v in stress_flows.items()},
        "total_programme_cost_m_per_yr": stress_costs,
    }

    # --- Module 3: self-financing break-even -------------------------------
    # Per-month fiscal value of one month less non-employment for a central
    # manufacturing worker re-employed at the post-displacement wage.
    # Convention: the marginal month is assumed to fall inside the first
    # 6 months of the claim, so the forgone-benefit saving = UC single-25+
    # standard allowance + NS JSA (weekly x 52/12).  See NOTES.md.
    post_wage_annual = (1 - p["penalty"]["central"]) * mean_wk * 52
    taxable = max(0.0, post_wage_annual - p["income_tax_personal_allowance"])
    nicable = max(0.0, post_wage_annual - p["nic_primary_threshold_annual"])
    tax_nic_annual = (taxable * p["income_tax_basic_rate"]
                      + nicable * p["nic_employee_main_rate"])
    benefit_month = (p["uc_single_25plus_monthly_2026_27"]
                     + p["jsa_weekly_2026_27"] * 52 / 12)
    fiscal_per_month = benefit_month + tax_nic_annual / 12
    wi_per_worker_yr = (p["wi_replacement"] * p["penalty"]["central"]
                        * mean_wk * 52)
    breakeven_months = wi_per_worker_yr / fiscal_per_month
    mods["breakeven"] = {
        "post_displacement_wage_annual": post_wage_annual,
        "income_tax_annual": taxable * p["income_tax_basic_rate"],
        "employee_nic_annual": nicable * p["nic_employee_main_rate"],
        "benefit_saving_per_month": benefit_month,
        "tax_nic_per_month": tax_nic_annual / 12,
        "fiscal_value_per_month": fiscal_per_month,
        "wage_insurance_per_worker_per_yr": wi_per_worker_yr,
        "breakeven_months_per_insured_worker": breakeven_months,
    }

    # --- Module 4: demographic composition of the eligible flow ------------
    # All-industry redundancies, latest year (Jul 2025 - Jun 2026); RED02/
    # RED01 do not split age or sex by industry (ASSUMPTION: manufacturing
    # composition proxied by all-industry).
    male_share = 100 * red01_latest_sex["men"] / red01_latest_sex["people"]
    fifty_share = 100 * red02_age["fifty_plus"] / red02_age["all"]
    mods["demographics"] = {
        "male_share_pct": male_share,
        "fifty_plus_share_pct": fifty_share,
        "basis": "all-industry redundancies, Jul 2025 - Jun 2026 "
                 "(RED01 NSA sex split; RED02 age bands)",
    }
    return mods


# --------------------------------------------------------------------------
# 5. LaTeX macro generation (\newcommand, prefix Taa; numbers never typed)
# --------------------------------------------------------------------------
def fmt(x, dp=0):
    return f"{x:,.{dp}f}"


def latex_macros(res, ashe, flows, red02_detail, mods):
    p = PARAMS
    m = []

    def add(name, value):
        m.append(f"\\newcommand{{\\Taa{name}}}{{{value}}}")

    def add_csname(name, value):
        # For macro names containing digits (requested grid naming
        # \TaaGridR30D1 etc.): LaTeX control words cannot contain digits,
        # so define and use these via \csname:
        #   \csname TaaGridR30D1\endcsname
        m.append(f"\\expandafter\\newcommand"
                 f"\\csname Taa{name}\\endcsname{{{value}}}")

    # Eligible flows (persons/yr)
    add("FlowManuf", fmt(flows["manuf"]))
    add("FlowAll", fmt(flows["all"]))
    add("FlowExposed", fmt(res["eligible_flows"]["exposed"]))
    add("FlowYearLabel", RED02_YEAR_LABEL)
    # Earnings (£/wk, FT gross, ASHE 2025 provisional)
    add("MeanWeeklyManuf", fmt(ashe["manuf"]["mean"], 2))
    add("MeanWeeklyAll", fmt(ashe["all"]["mean"], 2))
    add("MedianWeeklyManuf", fmt(ashe["manuf"]["median"], 2))
    add("MedianWeeklyAll", fmt(ashe["all"]["median"], 2))
    add("LowerQuartileWeeklyManuf", fmt(ashe["manuf"]["p25"], 2))
    add("LowerQuartileWeeklyAll", fmt(ashe["all"]["p25"], 2))
    # Statutory parameters (£)
    add("JsaWeekly", fmt(p["jsa_weekly_2026_27"], 2))
    add("UcSingleMonthly", fmt(p["uc_single_25plus_monthly_2026_27"], 2))
    add("UiWeekly", fmt(p["ui_weekly_2026_27"], 2))
    add("UiWeeklyGreenPaper", fmt(140.55, 2))  # 2025-26 rate quoted in GP era
    # Arm costs (£m/yr): \Taa<Arm>Cost<Scenario><Elig>
    arm_names = {"wage_insurance": "WageIns", "ui_extension": "UiExt",
                 "retraining": "Train", "uc_capital_disregard_bound": "UcDis"}
    for a, an in arm_names.items():
        for e in ["narrow", "broad", "exposed"]:
            for s in SCEN:
                add(f"{an}Cost{s.capitalize()}{e.capitalize()}",
                    fmt(res["arms"][a][e][s]))
    # Convenience central macros without eligibility suffix = narrow
    for a, an in arm_names.items():
        add(f"{an}CostCentral", fmt(res["arms"][a]["narrow"]["central"]))
    # Per-beneficiary (£/yr or £ one-off)
    for s in SCEN:
        add(f"WageInsPerWorker{s.capitalize()}Manuf",
            fmt(res["per_beneficiary"]["wage_insurance_annual_payment"][s]["narrow"]))
        add(f"WageInsPerWorker{s.capitalize()}All",
            fmt(res["per_beneficiary"]["wage_insurance_annual_payment"][s]["broad"]))
        add(f"UiExtPerWorker{s.capitalize()}",
            fmt(res["per_beneficiary"]["ui_extension_incremental"][s]))
    add("TrainPerWorker", fmt(p["train_stipend"]))
    add("UcDisPerWorker",
        fmt(res["per_beneficiary"]["uc_six_month_standard_allowance"]))
    # Totals and cross-checks
    for e in ["narrow", "broad", "exposed"]:
        for s in SCEN:
            add(f"TotalCost{s.capitalize()}{e.capitalize()}",
                fmt(res["totals"][e][s]))
    add("SteelSupportTotal", fmt(res["cross_checks"]["steel_support_m"]))
    add("SteelPortTalbot", fmt(p["steel_port_talbot_m"]))
    add("SteelScunthorpe", fmt(p["steel_scunthorpe_m"]))
    for e in ["narrow", "broad", "exposed"]:
        add(f"TotalCentralPctSteel{e.capitalize()}",
            fmt(res["cross_checks"]["total_central_pct_of_steel"][e]))
    # Quarterly RED02 detail (for the data appendix)
    for q, (tot, man) in red02_detail.items():
        tag = re.sub(r"[^A-Za-z]", "", q.split(" ")[0]) + q[-4:]
        # LaTeX macro names cannot contain digits -> spell the year
        tag = (tag[:-4] + {"2025": "TwentyFive", "2026": "TwentySix"}[tag[-4:]])
        add(f"RedManuf{tag}", fmt(man))
        add(f"RedAll{tag}", fmt(tot))

    # ---- Added analysis modules (appended; existing macros above are
    # ---- byte-identical to the previous version of this file) -------------
    m.append("% --- added analysis modules: wage-insurance grid, ISERBS, "
             "cyclical stress, break-even, demographics ---")
    # Module 1: 3x3 grid at the central (21.5%) penalty, £m/yr.  Digit-bearing
    # names as specified; use via \csname TaaGridR30D1\endcsname.
    for R in p["wi_grid_replacement"]:
        for D in p["wi_grid_duration_years"]:
            add_csname(f"GridR{int(R*100)}D{D}",
                       fmt(mods["wi_grid"]["cells_m_per_yr"]
                           [f"pencentral_r{int(R*100)}_d{D}"]))
    add("GridMin", fmt(mods["wi_grid"]["min_m_per_yr"]))
    add("GridMax", fmt(mods["wi_grid"]["max_m_per_yr"]))
    add("IserbsCost", fmt(mods["iserbs"]["cost_m_per_yr"]))
    add("IserbsPerWorker", fmt(mods["iserbs"]["per_worker_per_yr"]))
    # Module 2: cyclical stress (£m/yr, central narrow total programme)
    st = mods["stress"]
    add("CostPeakGFC", fmt(st["total_programme_cost_m_per_yr"]["gfc"]))
    add("CostPeakCovid", fmt(st["total_programme_cost_m_per_yr"]["covid"]))
    add("CostHistMedian",
        fmt(st["total_programme_cost_m_per_yr"]["hist_median"]))
    add("FlowRatioGFC", fmt(st["flow_ratio_to_current"]["gfc"], 1))
    # Module 3: self-financing break-even
    add("FiscalPerMonth", fmt(mods["breakeven"]["fiscal_value_per_month"]))
    add("BreakEvenMonths",
        fmt(mods["breakeven"]["breakeven_months_per_insured_worker"], 1))
    # Module 4: demographic composition (%, all-industry latest year)
    add("EligMaleShare", fmt(mods["demographics"]["male_share_pct"], 1))
    add("EligFiftyPlusShare",
        fmt(mods["demographics"]["fifty_plus_share_pct"], 1))

    header = ("% generated_taa.tex -- auto-generated by costing.py; do not edit.\n"
              "% All values trace to pinned raw files or labelled assumptions "
              "in NOTES.md.\n")
    return header + "\n".join(m) + "\n"


def main():
    verify_inputs()
    flow_all, flow_manuf, red02_detail = read_red02()
    flows = {"all": flow_all, "manuf": flow_manuf}
    ashe = read_ashe()
    res = cost_arms(flows, ashe)
    red01_annual, red01_latest_sex = read_red01_annual()
    red02_annual, red02_age = read_red02_history_and_age()
    mods = analysis_modules(res, flows, ashe, red01_annual, red01_latest_sex,
                            red02_annual, red02_age)
    res["inputs"] = {
        "redundancy_flows": {"annual_all": flow_all,
                             "annual_manufacturing": flow_manuf,
                             "quarters": {q: {"all": v[0], "manufacturing": v[1]}
                                          for q, v in red02_detail.items()},
                             "source": "ONS RED02 Aug 2026, Industry - levels"},
        "ashe_ft_gross_weekly_2025p": ashe,
        "parameters": {k: v for k, v in PARAMS.items()},
        "pinned_files_sha256": PINNED,
    }
    res["modules"] = mods
    res["inputs"]["red02_annual_history"] = {
        str(y): red02_annual[y] for y in sorted(red02_annual)}
    OUT.mkdir(exist_ok=True)
    (OUT / "results.json").write_text(json.dumps(res, indent=2))
    (OUT / "generated_taa.tex").write_text(
        latex_macros(res, ashe, flows, red02_detail, mods))
    print(f"Wrote {OUT/'results.json'} and {OUT/'generated_taa.tex'}")

    # Console summary
    print(f"\nEligible flows ({RED02_YEAR_LABEL}): "
          f"manufacturing {flow_manuf:,.0f}; all-industry {flow_all:,.0f}; "
          f"exposed memo {res['eligible_flows']['exposed']:,.0f}")
    print("\nHeadline costs, central scenario (GBP m/yr):")
    print(f"{'arm':32s}{'narrow':>10s}{'broad':>10s}{'exposed':>10s}")
    for a in res["arms"]:
        r = res["arms"][a]
        print(f"{a:32s}{r['narrow']['central']:>10,.0f}"
              f"{r['broad']['central']:>10,.0f}"
              f"{r['exposed']['central']:>10,.0f}")
    t = res["totals"]
    print(f"{'TOTAL':32s}{t['narrow']['central']:>10,.0f}"
          f"{t['broad']['central']:>10,.0f}{t['exposed']['central']:>10,.0f}")
    cc = res["cross_checks"]
    print(f"\nSteel support anchor: GBP {cc['steel_support_m']:.0f}m; "
          "central totals as % of it: "
          + ", ".join(f"{e} {v:.0f}%" for e, v
                      in cc["total_central_pct_of_steel"].items()))

    # Added modules
    g = mods["wi_grid"]["cells_m_per_yr"]
    print("\nWage-insurance grid at central 21.5% penalty (GBP m/yr):")
    print("        " + "".join(f"{d}yr".rjust(10) for d in (1, 2, 3)))
    for R in (30, 50, 70):
        print(f"  R{R}%  " + "".join(
            f"{g[f'pencentral_r{R}_d{d}']:>10,.0f}" for d in (1, 2, 3)))
    print(f"  27-cell grid min {mods['wi_grid']['min_m_per_yr']:,.0f}, "
          f"max {mods['wi_grid']['max_m_per_yr']:,.0f}")
    i = mods["iserbs"]
    print(f"ISERBS cell (90% top-up, 78wk): GBP {i['cost_m_per_yr']:,.0f}m/yr"
          f"; per worker GBP {i['per_worker_per_yr']:,.0f}/yr")
    st = mods["stress"]
    sc = st["total_programme_cost_m_per_yr"]
    print(f"Stress totals (central narrow): GFC {sc['gfc']:,.0f}m "
          f"(flow x{st['flow_ratio_to_current']['gfc']:.1f}), "
          f"Covid {sc['covid']:,.0f}m, hist median {sc['hist_median']:,.0f}m")
    b = mods["breakeven"]
    print(f"Break-even: fiscal value GBP {b['fiscal_value_per_month']:,.0f}"
          f"/month -> {b['breakeven_months_per_insured_worker']:.1f} months "
          "of reduced non-employment per insured worker")
    d = mods["demographics"]
    print(f"Eligible-flow composition (all-industry): male "
          f"{d['male_share_pct']:.1f}%, 50+ {d['fifty_plus_share_pct']:.1f}%")


if __name__ == "__main__":
    main()
